
import openpyxl
import os
import shutil


# set up "input.xlsx" for reading
input_file = "C:\\Users\\Peyton.Dexter\\Highstreet Work\\Resume Copy\\For upload\\input_clean.xlsx"
wb = openpyxl.load_workbook(filename=input_file)
ws = wb.active


# set up source and dest
source = "C:\\Users\\Peyton.Dexter\\Highstreet Work\\Resume Copy\\Backup_Resume_1\\"
dest = "C:\\Users\\Peyton.Dexter\\Highstreet Work\\Resume Copy\\For upload\\Resumes\\"

# for the entire spreadsheet, if name exists in source, copy to destination
for row in ws.iter_rows(min_row=0, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    print(row[7].value)

    if row[7].value is not None:
        if os.path.isfile(source + row[7].value):
            shutil.copy(source + row[7].value, dest + row[7].value)
            print("Copied\n")

            
